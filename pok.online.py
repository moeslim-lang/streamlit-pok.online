import io
import re
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfgen import canvas as pdfcanvas

st.set_page_config(page_title="Dashboard Anggaran", layout="wide")

st.title("📊 Dashboard Data Anggaran TA 2025")

# =========================== UPLOAD FILE =======================================
uploaded = st.sidebar.file_uploader("Upload File Excel", type=["xlsx"])
if not uploaded:
    st.stop()

# Baca workbook untuk dapat memilih sheet
xls = pd.ExcelFile(uploaded)
sheets = xls.sheet_names
sheet_selected = st.sidebar.selectbox("Pilih Sheet", sheets)

# Muat sheet yang dipilih
df = pd.read_excel(xls, sheet_name=sheet_selected, dtype=str)

# ======================== PERBAIKAN DATA DASAR ================================
# Kolom wajib
wajib = ["UNIT","MAK","KODE","URAIAN","VOL","SAT","HARGA","JUMLAH","RO","SD"]
df = df.reindex(columns=wajib)

# Bersihkan None dan NaN
df = df.replace(["None", "none", "NONE"], "")
df = df.fillna("")

# Konversi angka
for col in ["VOL","HARGA","JUMLAH"]:
    df[col] = pd.to_numeric(df[col], errors="coerce")

# ========================= FILTER / NAV ======================================
st.sidebar.header("Filter")
units = df["UNIT"].replace("", pd.NA).dropna().unique().tolist()
# Default view: show summary (rekap) per unit. User can then pick a unit to
# view rincian. Provide choices: 'Rekap Per Unit', 'Semua', plus individual units.
unit_selected = st.sidebar.selectbox("Pilih Unit", ["Rekap Per Unit", "Semua"] + units)

# If the user chose Rekap, we will show the grouped summary and let them pick
# a unit from the main area to see details. Otherwise, filter immediately.
df_filtered = df.copy()
detail_unit = None
if unit_selected == "Rekap Per Unit":
    # keep df_filtered as full dataset for the rekap view
    pass
elif unit_selected == "Semua":
    df_filtered = df.copy()
else:
    # specific unit chosen in sidebar -> show details for that unit
    df_filtered = df[df["UNIT"] == unit_selected]
    detail_unit = unit_selected

# ========================= HITUNG TOTAL AKUN (6 DIGIT) ========================
mask_akun = df_filtered["KODE"].astype(str).str.match(r"^\d{6}$")
total_anggaran = df_filtered.loc[mask_akun, "JUMLAH"].sum()

total_fmt = f"{total_anggaran:,.0f}".replace(",", ".")

st.metric("Total Anggaran (Kode Akun 6 Digit)", total_fmt)

# ========================= TAMPILKAN TABEL ====================================
df_display = df_filtered.copy()

# Pastikan kolom VOL/HARGA/JUMLAH tetap sebagai numeric sehingga Streamlit
# akan merendernya rata kanan. Gunakan Styler.format untuk menampilkan
# angka dengan pemisah ribuan (titik) tetapi pertahankan tipe aslinya.
df_display["VOL"] = pd.to_numeric(df_display["VOL"], errors="coerce")
df_display["HARGA"] = pd.to_numeric(df_display["HARGA"], errors="coerce")
df_display["JUMLAH"] = pd.to_numeric(df_display["JUMLAH"], errors="coerce")

# If we're in Rekap view, show summary per `UNIT` first; allow user to
# pick a unit to view rincian. Otherwise show the detail table for the
# currently-filtered dataset.
if unit_selected == "Rekap Per Unit" and detail_unit is None:
    st.subheader("Rekap Per Unit")
    # Compute summary: total JUMLAH (for 6-digit kode?) and count rows
    # We'll show total anggaran (all JUMLAH) and count of records for each UNIT.
    summary = df.copy()
    summary["JUMLAH"] = pd.to_numeric(summary["JUMLAH"], errors="coerce").fillna(0)
    # Mark rows where KODE is exactly 6 digits
    summary["is6"] = summary["KODE"].astype(str).str.fullmatch(r"\d{6}")

    # Group only rows with is6==True to compute counts and totals per UNIT
    grp6 = summary[summary["is6"]].groupby("UNIT", dropna=False).agg(Count=("KODE", "count"), Total_JUMLAH=("JUMLAH", "sum")).reset_index()

    # Ensure all units appear in the rekap (even those with zero matching KODE)
    all_units = summary["UNIT"].fillna("").unique().tolist()
    grp = pd.DataFrame({"UNIT": all_units}).merge(grp6, on="UNIT", how="left").fillna({"Count": 0, "Total_JUMLAH": 0})
    grp["Count"] = grp["Count"].astype(int)
    grp = grp.reset_index(drop=True)

    grp["Total_JUMLAH_fmt"] = grp["Total_JUMLAH"].apply(lambda x: "" if x == 0 else f"{x:,.0f}".replace(",", "."))

    # Display only UNIT and Total JUMLAH (no percent)
    st.dataframe(grp[["UNIT", "Total_JUMLAH_fmt"]].rename(columns={"Total_JUMLAH_fmt":"Total JUMLAH"}), use_container_width=True)

    # Provide download for the rekap as Excel (UNIT and Total JUMLAH)
    try:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            export_df = grp[["UNIT", "Total_JUMLAH"]].copy()
            export_df.rename(columns={"Total_JUMLAH":"Total JUMLAH"}, inplace=True)
            export_df.to_excel(writer, index=False, sheet_name="Rekap Per Unit")
        buf.seek(0)
        st.download_button("⬇ Download Rekap Excel", buf.getvalue(), file_name="Rekap_Per_Unit.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Also prepare PDF version of the rekap (landscape, centered title)
        try:
            pdf_buf = io.BytesIO()
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet

            doc = SimpleDocTemplate(pdf_buf, pagesize=landscape(A4), leftMargin=12*mm, rightMargin=12*mm, topMargin=12*mm, bottomMargin=12*mm)
            styles = getSampleStyleSheet()
            elems = []
            title_text = f"RINCIAN KERTAS KERJA SATKER T.A. 2025 ({sheet_selected})"
            title_style = ParagraphStyle(name='rekap_title', parent=styles['Heading2'], alignment=1)
            elems.append(Paragraph(title_text, title_style))
            elems.append(Spacer(1, 6))

            # Build table data with UNIT and Total JUMLAH only
            pdf_rows = [["UNIT", "Total JUMLAH"]]
            for _, r in grp.iterrows():
                totalj = "" if r['Total_JUMLAH'] == 0 else f"{int(r['Total_JUMLAH']):,}".replace(",", ".")
                pdf_rows.append([r['UNIT'], totalj])

            # Wider columns for landscape layout
            t = Table(pdf_rows, colWidths=[160*mm, 50*mm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ('ALIGN', (1,1), (1,-1), 'RIGHT'),
            ]))
            elems.append(t)
            doc.build(elems)
            pdf_buf.seek(0)
            st.download_button("⬇ Download Rekap PDF", pdf_buf.getvalue(), file_name="Rekap_Per_Unit.pdf", mime="application/pdf")
        except Exception:
            pass

    except Exception:
        # If Excel export fails, offer CSV as fallback (UNIT, Total_JUMLAH)
        csv_buf = grp[["UNIT", "Total_JUMLAH"]].to_csv(index=False).encode("utf-8")
        st.download_button("⬇ Download Rekap CSV", csv_buf, file_name="Rekap_Per_Unit.csv", mime="text/csv")

    # Inform user to pick a unit from the sidebar to view rincian
    st.info("Pilih unit dari sidebar untuk melihat rincian")
    st.stop()

st.subheader("Tabel Data")
# Render tabel sebagai HTML dengan CSS untuk memastikan kolom VOL/HARGA/JUMLAH
# diratakan ke kanan (beberapa versi Streamlit tidak merender Styler CSS).
try:
    df_html = df_display.copy()
    # Format angka menjadi string tampilan (titik sebagai pemisah ribuan)
    df_html["VOL"] = df_html["VOL"].apply(lambda x: "" if pd.isna(x) or x == 0 else f"{int(x):,}".replace(",", "."))
    df_html["HARGA"] = df_html["HARGA"].apply(lambda x: "" if pd.isna(x) or x == 0 else f"{x:,.0f}".replace(",", "."))
    df_html["JUMLAH"] = df_html["JUMLAH"].apply(lambda x: "" if pd.isna(x) or x == 0 else f"{x:,.0f}".replace(",", "."))

    cols = list(df_html.columns)
    # Build CSS: sticky header, spacing, alignment for numeric columns, and our conditional classes
    css_parts = [
        "<style>",
        "table.dataframe{border-collapse:collapse;width:100%;font-family:Arial,Helvetica,sans-serif}",
        "table.dataframe td, table.dataframe th{padding:2px 6px;font-size:11px;line-height:1.1}",
        "table.dataframe thead th{position:sticky;top:0;background:#E0E0E0;z-index:3;text-align:center}",
        # classes for rows where KODE is text, numeric, or 6-digit
        ".kode-text td{background:#CCE5FF}",
        ".kode-numeric td.numeric{font-weight:700;font-style:italic}",
        ".kode-6digit td{font-weight:700;font-style:italic}",
    ]

    # Add right alignment for VOL/HARGA/JUMLAH if present
    for colname in ("VOL", "HARGA", "JUMLAH"):
        if colname in cols:
            idx = cols.index(colname) + 1
            css_parts.append(f"table.dataframe td:nth-child({idx}){{text-align:right}}")

    css_parts.append("</style>")
    css = "".join(css_parts)

    # Build HTML table manually so we can add row classes based on KODE value
    header_cells = "".join([f"<th>{c}</th>" for c in cols])
    rows_html = []
    for _, row in df_html.iterrows():
        kode_val = "" if pd.isna(row.get("KODE", "")) else str(row.get("KODE", "")).strip()
        # Determine class based on KODE only
        if kode_val == "":
            row_class = ""
        elif re.search(r"[A-Za-z]", kode_val):
            row_class = "kode-text"
        elif re.fullmatch(r"\d{6}", kode_val):
            # KODE consists of exactly 6 digits -> bold + italic entire row
            row_class = "kode-6digit"
        elif re.fullmatch(r"\d+", kode_val):
            # other numeric KODE values (not 6 digits)
            row_class = "kode-numeric"
        else:
            row_class = "kode-text"

        # Build cells; mark numeric columns with class 'numeric' so CSS can target them
        cell_html = []
        for c in cols:
            v = row.get(c, "")
            cell_value = "" if pd.isna(v) else str(v)
            td_class = "numeric" if c in {"VOL", "HARGA", "JUMLAH"} else ""
            if td_class:
                cell_html.append(f"<td class='{td_class}'>{cell_value}</td>")
            else:
                cell_html.append(f"<td>{cell_value}</td>")

        tr = f"<tr class='{row_class}'>" + "".join(cell_html) + "</tr>"
        rows_html.append(tr)

    table_html = f"<table class='dataframe'><thead><tr>{header_cells}</tr></thead><tbody>{''.join(rows_html)}</tbody></table>"
    container_html = f'<div style="max-height:400px;overflow:auto">{table_html}</div>'
    st.markdown(css + container_html, unsafe_allow_html=True)
except Exception:
    # Fallback: tampilkan DataFrame biasa
    st.dataframe(df_display, use_container_width=True, hide_index=True)

# ========================== EXPORT EXCEL ======================================
def generate_excel(dataframe):
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rincian"

    # Header
    header_fill = PatternFill(start_color="E0E0E0", fill_type="solid")
    for i, col in enumerate(dataframe.columns, start=1):
        c = ws.cell(row=1, column=i, value=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = header_fill


    # Data with conditional formatting rules based ONLY on the KODE column:
    # - If KODE contains alphabetic characters (text) -> fill entire row light blue.
    # - If KODE is numeric (digits) -> for that row, numeric cells are bold+italic.
    # - If KODE is empty -> leave row normal.
    row_text_fill = PatternFill(start_color="CCE5FF", fill_type="solid")

    # find index of KODE column (0-based)
    try:
        kode_idx = list(dataframe.columns).index("KODE")
    except ValueError:
        kode_idx = None

    for r, row in enumerate(dataframe.itertuples(index=False), start=2):
        # Determine KODE value for this row
        kode_val = None
        if kode_idx is not None:
            try:
                kode_val = row[kode_idx]
            except Exception:
                kode_val = None

        kode_str = "" if kode_val is None else str(kode_val).strip()

        row_is_text = False
        row_is_numeric = False
        if kode_str == "":
            # empty -> normal
            pass
        elif re.search(r"[A-Za-z]", kode_str):
            row_is_text = True
        elif re.fullmatch(r"\d+", kode_str):
            row_is_numeric = True
        else:
            # fallback: if it isn't purely digits but also not alphabetic, treat as text
            row_is_text = True

        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # If KODE indicates text, fill entire row
            if row_is_text:
                cell.fill = row_text_fill

            # If KODE indicates numeric, style numeric cells bold+italic
            if row_is_numeric:
                is_numeric = False
                if isinstance(val, (int, float)):
                    is_numeric = True
                elif isinstance(val, str):
                    s = val.strip()
                    if re.fullmatch(r"-?\d+(?:\.\d+)?", s):
                        is_numeric = True
                if is_numeric:
                    cell.font = Font(bold=True, italic=True)

            # Preserve / add KODE 6-digit highlight (overrides other fills on that cell)
            if dataframe.columns[c-1] == "KODE":
                if isinstance(kode_val, str) and re.fullmatch(r"\d{6}", kode_val):
                    cell.fill = PatternFill(start_color="CCE5FF", fill_type="solid")

        # Adjust row height (openpyxl uses points)
        try:
            ws.row_dimensions[r].height = 13
        except Exception:
            pass

    wb.save(buffer)
    return buffer.getvalue()

if detail_unit is not None or unit_selected != "Rekap Per Unit":
    xlsx = generate_excel(df_display)
else:
    xlsx = None

# =========================== EXPORT PDF ========================================
def generate_pdf(dataframe, sheet_name, unit_name, total_anggaran, include_signature=True):

    # Buat PDF landscape A4 dengan margin; tambahkan bottom margin lebih besar
    # agar blok tanda tangan tidak tertimpa tabel.
    buffer = io.BytesIO()
    left_margin = right_margin = top_margin = 12 * mm
    # Reduce bottom margin to allow more table rows per page while still
    # leaving space for the signature block. 18 mm is a reasonable compromise.
    bottom_margin = 18 * mm
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=left_margin, rightMargin=right_margin,
                            topMargin=top_margin, bottomMargin=bottom_margin)

    styles = getSampleStyleSheet()
    title_style = styles['Heading2']
    title_style.alignment = 1  # center
    title_style.fontSize = 11
    title_style.spaceAfter = 6

    info_style = styles['Normal']
    info_style.fontSize = 9

    title_text = f"RINCIAN KERTAS KERJA SATKER T.A. 2025 ({sheet_name})"
    total_fmt = f"{total_anggaran:,.0f}".replace(",", ".")
    unit_text = f"Unit: {unit_name}                                        Total: {total_fmt}"

    elems = []
    elems.append(Paragraph(title_text, title_style))
    elems.append(Spacer(1, 4))
    elems.append(Paragraph(unit_text, info_style))
    elems.append(Spacer(1, 6))

    # Siapkan data tabel dengan lebar kolom tetap (dalam mm). Jika total melebihi
    # lebar tersedia, ukurannya akan diskalakan secara proporsional.
    page_width, page_height = landscape(A4)
    avail_width = page_width - left_margin - right_margin

    # Default lebar tiap kolom dalam mm (sesuaikan bila perlu)
    fixed_widths_mm = {
        'UNIT': 33,
        'MAK': 35,
        'KODE': 19,
        'URAIAN': 150,
        'VOL': 15,
        'SAT': 15,
        'HARGA': 27,
        'JUMLAH': 30,
        'RO': 7,
        'SD': 10,
    }

    # Bangun list lebar (dalam points)
    fixed_points = []
    cols = list(dataframe.columns)
    for col in cols:
        mm_val = fixed_widths_mm.get(str(col).upper(), 25)
        fixed_points.append(mm_val * mm)

    total_fixed = sum(fixed_points)
    if total_fixed > avail_width and total_fixed > 0:
        scale = avail_width / total_fixed
        col_widths = [w * scale for w in fixed_points]
    else:
        col_widths = fixed_points

    # Siapkan style untuk sel tabel (pakai Paragraph untuk kolom URAIAN agar wrapping bekerja)
    styles = getSampleStyleSheet()
    cell_style = styles['BodyText']
    # Slightly larger font and leading so rows are not too cramped
    cell_style.fontSize = 8
    cell_style.leading = 10
    cell_style.alignment = 0  # left

    # Header + data: create Paragraphs for every data cell so we can reliably
    # control font per-row. Also compute KODE classification for each row.
    header = cols
    rows = []
    kode_row_classes = []
    for row in dataframe.values.tolist():
        # Determine KODE for this row first
        try:
            kode_val = row[cols.index('KODE')]
        except Exception:
            kode_val = ''
        kode_str = '' if pd.isna(kode_val) else str(kode_val).strip()
        if kode_str == '':
            kode_class = 'empty'
        elif re.search(r'[A-Za-z]', kode_str):
            kode_class = 'text'
        elif re.fullmatch(r'\d{6}', kode_str):
            kode_class = 'six'
        elif re.fullmatch(r'\d+', kode_str):
            kode_class = 'numeric'
        else:
            kode_class = 'text'
        kode_row_classes.append(kode_class)

        # Build a per-row ParagraphStyle derived from cell_style
        # base style
        base_font = 'Helvetica'
        bold_italic_font = 'Helvetica-BoldOblique'
        if kode_class == 'six':
            row_font = bold_italic_font
        else:
            row_font = base_font

        new_row = []
        for i, val in enumerate(row):
            colname = cols[i]
            cname = str(colname).upper()
            if cname == 'VOL':
                s = '' if pd.isna(val) or val == 0 else f"{int(val):,}".replace(',', '.')
            elif cname in {'HARGA', 'JUMLAH'}:
                s = '' if pd.isna(val) or val == 0 else f"{val:,.0f}".replace(',', '.')
            else:
                s = '' if pd.isna(val) else str(val)

            # create style per cell so alignment can be adjusted for numeric columns
            if cname in {'VOL', 'HARGA', 'JUMLAH'}:
                pstyle = ParagraphStyle(name=f'ps_{kode_class}_{cname}', parent=cell_style,
                                        fontName=row_font, alignment=TA_RIGHT)
            else:
                pstyle = ParagraphStyle(name=f'ps_{kode_class}_{cname}', parent=cell_style,
                                        fontName=row_font, alignment=0)

            new_row.append(Paragraph(s, pstyle))

        rows.append(new_row)

    data_table = [header] + rows

    # Let ReportLab compute row heights automatically so text wrapping is
    # accommodated. This prevents truncation or overly tight rows when URAIAN
    # wraps across lines.
    table = Table(data_table, colWidths=col_widths, repeatRows=1)

    # Styling: header, font size, alignment per kolom, vertical middle for URAIAN
    ts = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        # Use slightly larger padding so wrapped text has room
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])

    # Right align numeric-like columns
    numeric_cols = [i for i, c in enumerate(cols) if str(c).upper() in {"VOL", "HARGA", "JUMLAH"}]
    for ci in numeric_cols:
        ts.add('ALIGN', (ci, 1), (ci, -1), 'RIGHT')

    # Apply per-row styling based on KODE classification
    # data rows in the ReportLab table start at row index 1 (header is row 0)
    for i, cls in enumerate(kode_row_classes):
        table_row = i + 1
        if cls == 'text':
            # light blue background across the entire row
            try:
                ts.add('BACKGROUND', (0, table_row), (-1, table_row), colors.HexColor('#CCE5FF'))
            except Exception:
                # fallback to lightblue constant
                ts.add('BACKGROUND', (0, table_row), (-1, table_row), colors.lightblue)
        elif cls == 'six':
            # entire row bold + italic
            # use Helvetica-BoldOblique if available
            ts.add('FONTNAME', (0, table_row), (-1, table_row), 'Helvetica-BoldOblique')

    # URAIAN column: ensure left align and vertical middle
    try:
        ur_idx = cols.index('URAIAN')
        ts.add('ALIGN', (ur_idx, 1), (ur_idx, -1), 'LEFT')
        ts.add('VALIGN', (ur_idx, 0), (ur_idx, -1), 'MIDDLE')
    except ValueError:
        pass

    table.setStyle(ts)

    elems.append(table)

    # Custom canvas to add page numbers and signature block on the last page
    class NumberedCanvas(pdfcanvas.Canvas):
        pass

    def _make_numbered_canvas(left_margin, right_margin, top_margin, bottom_margin, include_signature):
        class NC(pdfcanvas.Canvas):
            def __init__(self, *args, **kwargs):
                pdfcanvas.Canvas.__init__(self, *args, **kwargs)
                self._saved_page_states = []

            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                num_pages = len(self._saved_page_states)
                for i, state in enumerate(self._saved_page_states):
                    self.__dict__.update(state)
                    # draw page number centered at bottom
                    page_num = i + 1
                    page_text = f"{page_num} / {num_pages}"
                    pw, ph = self._pagesize
                    x_center = pw / 2.0
                    y_page = bottom_margin * 0.3
                    self.setFont('Helvetica', 8)
                    self.drawCentredString(x_center, y_page, page_text)

                    # If last page, optionally draw signature block at bottom-right
                    if page_num == num_pages and include_signature:
                        # signature lines (right-aligned)
                        sig_lines = [
                            "Lhokseumawe, 03 Oktober 2025",
                            "Wakil Rektor II",
                            "Bidang Administrasi Umum, Perencanaan, dan Keuangan",
                            "",
                            "",
                            "^",
                            "",
                            "",
                            "SAID ALWI",
                        ]
                        self.setFont('Helvetica', 9)
                        x_left = left_margin
                        leading = 12
                        # Position so the last signature line is about 3 lines above bottom margin
                        last_line_y = bottom_margin + (leading * 3)
                        first_line_y = last_line_y + leading * (len(sig_lines) - 1)
                        # ensure not beyond top margin
                        max_top = ph - top_margin - 6
                        if first_line_y > max_top:
                            first_line_y = max_top
                        # draw lines top-down, left-aligned
                        for idx, line in enumerate(sig_lines):
                            y = first_line_y - idx * leading
                            if line.strip() == "^":
                                # draw caret slightly indented
                                self.drawString(x_left + 6 * mm, y, line)
                            else:
                                self.drawString(x_left, y, line)

                    pdfcanvas.Canvas.showPage(self)
                pdfcanvas.Canvas.save(self)

        return NC

    CanvasMaker = _make_numbered_canvas(left_margin, right_margin, top_margin, bottom_margin, include_signature)
    doc.build(elems, canvasmaker=CanvasMaker)
    return buffer.getvalue()

# PDF generation controlled by sidebar checkbox (include signature)
include_signature = st.sidebar.checkbox("Sertakan tanda tangan pada PDF", value=True)
if xlsx is not None:
    pdf = generate_pdf(df_display, sheet_selected, unit_selected if detail_unit is None else detail_unit, total_anggaran, include_signature)
else:
    pdf = None

# ========================= BUTTON DOWNLOAD ====================================
st.sidebar.header("Export")
if xlsx is not None:
    st.sidebar.download_button("⬇ Download Excel", xlsx,
        file_name=f"Rincian_{detail_unit or unit_selected}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if pdf is not None:
    st.sidebar.download_button("⬇ Download PDF (Landscape)", pdf,
        file_name=f"Rincian_{detail_unit or unit_selected}.pdf",
        mime="application/pdf")

# Sidebar contact pinned to bottom
contact_html = '''
<div style="position:fixed; bottom:12px; left:12px; width:220px;">
    <hr style="margin:6px 0;">
    <div style="font-size:12px;color:#374151">cp: <a href="mailto:muslim@uinsuna.ac.id">muslim@uinsuna.ac.id</a></div>
</div>
'''
st.sidebar.markdown(contact_html, unsafe_allow_html=True)

